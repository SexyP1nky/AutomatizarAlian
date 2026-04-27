"""
Responsabilidade: montar o arquivo XLSX de saída com as regras definidas.

Estrutura do arquivo gerado:
  Aba "ESTORNOS":
    Linha 1  (A1)  : Mês/ano atual + "VENDEDORES", em maiúsculas, negrito
    Linha 2        : Cabeçalhos em negrito/maiúsculas: CLIENTE | SITUAÇÃO | COMISSAO | VEND/CORRETORA | VALOR ESTORNO | OBSERVAÇÃO
    Linhas 3..N   : Uma linha por MatchedRecord
    Última linha  : TOTAL destacado com fundo amarelo na coluna VALOR ESTORNO

  Aba "NÃO ENCONTRADOS" (quando houver):
    Linha 1        : Cabeçalho "SEGURADO" em negrito
    Linhas 2..N   : Um nome por linha

Regra de VALOR ESTORNO:
  - Conta quantas vezes cada vendedor aparece na tabela de saída (coluna VEND/CORRETORA)
  - 1 a 3 aparições → R$ 30,00 por linha
  - 4 ou mais aparições → R$ 50,00 por linha
"""

from collections import Counter
from datetime import datetime
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from matcher import MatchedRecord
from vendor_sales_counter import get_estorno_value


# ── Constantes de formatação ──────────────────────────────────────────────────

_FONT_NAME = "Arial"
_COL_HEADERS = ["CLIENTE", "SITUAÇÃO", "COMISSAO", "VEND/CORRETORA", "Nº APÓLICE", "VALOR ESTORNO", "OBSERVAÇÃO"]
_COL_WIDTHS = [42, 12, 14, 22, 18, 16, 40]

_FILL_HEADER = PatternFill("solid", start_color="D9D9D9")   # cinza claro
_FILL_TOTAL = PatternFill("solid", start_color="FFFF00")     # amarelo
_FILL_WARNING = PatternFill("solid", start_color="FFCC99")   # laranja claro para alertas

_MONTHS_PT = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR",
    5: "MAI", 6: "JUN", 7: "JUL", 8: "AGO",
    9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}


# ── Funções auxiliares ────────────────────────────────────────────────────────

def _format_inicio_vig(date_str: str) -> str:
    """
    Converte DD/MM/YYYY → ABR/2026.
    Retorna a string original em caso de formato inesperado.
    """
    try:
        parts = date_str.strip().split("/")
        if len(parts) == 3:
            month = int(parts[1])
            year = parts[2]
            return f"{_MONTHS_PT[month]}/{year}"
    except (ValueError, KeyError, IndexError):
        pass
    return date_str


def _current_month_label() -> str:
    """Retorna ex: 'ABR/26 VENDEDORES'"""
    now = datetime.now()
    year_short = str(now.year)[2:]
    return f"{_MONTHS_PT[now.month]}/{year_short} VENDEDORES"


def _calculate_vendor_values(
    records: List[MatchedRecord],
    vendor_sales_counts: Dict[Tuple[str, str], int],
) -> List[int]:
    """
    Calcula o valor de estorno (30 ou 50) para cada registro.

    Usa as contagens de vendas positivas por vendedor/mês para determinar:
      - 4+ vendas positivas naquele mês → R$ 50
      - 3 ou menos → R$ 30

    Retorna lista de valores na mesma ordem dos records.
    """
    return [
        get_estorno_value(vendor_sales_counts, r.vendedor, r.inicio_vig)
        for r in records
    ]


# ── Funções de escrita de células ─────────────────────────────────────────────

def _write_title_row(ws, label: str) -> None:
    cell = ws["A1"]
    cell.value = label
    cell.font = Font(name=_FONT_NAME, bold=True, size=12)
    cell.alignment = Alignment(horizontal="left")


def _write_column_headers(ws) -> None:
    for col_idx, header in enumerate(_COL_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name=_FONT_NAME, bold=True)
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")


def _write_data_rows(
    ws, records: List[MatchedRecord], vendor_values: List[int]
) -> None:
    for row_idx, record in enumerate(records, start=3):
        valor = vendor_values[row_idx - 3]

        # Dados normais
        ws.cell(row=row_idx, column=1, value=record.segurado.upper())
        ws.cell(row=row_idx, column=2, value="ESTORNO")
        ws.cell(row=row_idx, column=3, value=record.sheet_name.upper())
        ws.cell(row=row_idx, column=4, value=record.vendedor.upper())
        ws.cell(row=row_idx, column=5, value=record.apolice_pdf)

        valor_cell = ws.cell(row=row_idx, column=6, value=valor)
        valor_cell.number_format = 'R$ #,##0'

        # Tratamento de OBSERVAÇÃO para matches incertos
        obs_text = ""
        if record.match_type != "EXATO":
            if record.match_type == "APOLICE_DIFERENTE":
                obs_text = f"VERIFICAR - APÓLICE DIFERENTE (PDF: {record.apolice_pdf} vs XLSX: {record.apolice_xlsx})"
            elif record.match_type == "FUZZY_APOLICE_DIFERENTE":
                obs_text = f"VERIFICAR - MATCH APROXIMADO E APÓLICE DIFERENTE (PDF: {record.apolice_pdf} vs XLSX: {record.apolice_xlsx})"
            elif record.match_type == "FUZZY":
                obs_text = "VERIFICAR - MATCH APROXIMADO PELO NOME"
            elif record.match_type == "APOLICE_AUSENTE_XLSX":
                obs_text = "VERIFICAR - NÚMERO DE APÓLICE AUSENTE NA PLANILHA"
            elif record.match_type == "FUZZY_APOLICE_AUSENTE_XLSX":
                obs_text = "VERIFICAR - MATCH APROXIMADO E NÚMERO DE APÓLICE AUSENTE NA PLANILHA"
                
            obs_cell = ws.cell(row=row_idx, column=7, value=obs_text)
            
            # Pinta a linha inteira de laranja claro para chamar atenção
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = _FILL_WARNING


def _write_total_row(ws, total_row: int, first_data_row: int) -> None:
    bold_font = Font(name=_FONT_NAME, bold=True)

    label_cell = ws.cell(row=total_row, column=4, value="TOTAL")
    label_cell.font = bold_font
    label_cell.fill = _FILL_TOTAL
    label_cell.alignment = Alignment(horizontal="center")

    # Fórmula SUM para que o total recalcule se o usuário editar valores
    last_data_row = total_row - 1
    total_cell = ws.cell(
        row=total_row,
        column=6,
        value=f"=SUM(F{first_data_row}:F{last_data_row})",
    )
    total_cell.font = bold_font
    total_cell.fill = _FILL_TOTAL
    total_cell.number_format = 'R$ #,##0'


def _set_column_widths(ws) -> None:
    col_letters = ["A", "B", "C", "D", "E", "F", "G"]
    for letter, width in zip(col_letters, _COL_WIDTHS):
        ws.column_dimensions[letter].width = width


def _write_not_found_sheet(wb, not_found: List[str]) -> None:
    """
    Cria uma aba "NÃO ENCONTRADOS" no workbook com a lista de segurados
    que não foram localizados na planilha de comissões.
    """
    ws = wb.create_sheet(title="NÃO ENCONTRADOS")

    # Cabeçalho
    header_cell = ws.cell(row=1, column=1, value="SEGURADO")
    header_cell.font = Font(name=_FONT_NAME, bold=True)
    header_cell.fill = _FILL_HEADER
    header_cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, name in enumerate(not_found, start=2):
        cell = ws.cell(row=row_idx, column=1, value=name.upper())
        cell.font = Font(name=_FONT_NAME)

    # Largura da coluna
    ws.column_dimensions["A"].width = 50


# ── Ponto de entrada público ──────────────────────────────────────────────────

def build_report(
    records: List[MatchedRecord],
    output_path: str,
    not_found: List[str] | None = None,
    vendor_sales_counts: Dict[Tuple[str, str], int] | None = None,
) -> None:
    """
    Gera o XLSX de controle de estornos em output_path.
    Levanta ValueError se records estiver vazio.

    Parâmetros:
        records              — lista de MatchedRecord (segurado × vendedor)
        output_path          — caminho do arquivo de saída
        not_found            — lista de nomes de segurados sem correspondência (opcional)
        vendor_sales_counts  — contagens de vendas positivas por (vendedor, mês)
    """
    if not records:
        raise ValueError("Nenhum registro correspondido — relatório não gerado.")

    if vendor_sales_counts is None:
        vendor_sales_counts = {}

    vendor_values = _calculate_vendor_values(records, vendor_sales_counts)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ESTORNOS"

    _write_title_row(ws, _current_month_label())
    _write_column_headers(ws)

    first_data_row = 3
    _write_data_rows(ws, records, vendor_values)

    total_row = first_data_row + len(records)
    _write_total_row(ws, total_row, first_data_row)

    _set_column_widths(ws)

    # Aba de não encontrados (quando houver)
    if not_found:
        _write_not_found_sheet(wb, not_found)

    wb.save(output_path)
